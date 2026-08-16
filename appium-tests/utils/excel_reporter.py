"""
Excel Report Analysis Generator for Appium Mobile E2E Test Suite
Uses openpyxl to generate styled spreadsheets with charts, KPI summary blocks, and test diagnostics.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from datetime import datetime
from utils.logger import logger


class ExcelAnalysisReporter:
    def __init__(self, output_dir=None):
        if not output_dir:
            output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")
        os.makedirs(output_dir, exist_ok=True)
        self.output_dir = output_dir
        self.wb = openpyxl.Workbook()

    def generate_report(self, test_results, total_duration_sec):
        """
        Generates a comprehensive Excel report workbook (.xlsx)
        :param test_results: list of dicts containing test outcome details
        :param total_duration_sec: total execution time in seconds
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Appium_Android_E2E_Test_Report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        # ----------------------------------------------------
        # 1. Executive Summary Sheet
        # ----------------------------------------------------
        ws_summary = self.wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        total_tests = len(test_results)
        passed_tests = sum(1 for r in test_results if r.get('status') == 'PASS')
        failed_tests = sum(1 for r in test_results if r.get('status') == 'FAIL')
        skipped_tests = sum(1 for r in test_results if r.get('status') == 'SKIPPED')
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0

        # Header Title Banner
        ws_summary.merge_cells("A1:G2")
        title_cell = ws_summary["A1"]
        title_cell.value = "📱 APPIUM ANDROID END-TO-END AUTOMATION ANALYSIS REPORT"
        title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Metadata Sub-header
        ws_summary.merge_cells("A3:G3")
        meta_cell = ws_summary["A3"]
        meta_cell.value = f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: Android Chrome/Native | Total Duration: {total_duration_sec:.2f}s"
        meta_cell.font = Font(name="Segoe UI", size=10, italic=True, color="64748B")
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")

        # KPI Metrics Cards (Rows 5 to 7)
        kpis = [
            ("TOTAL TESTS", total_tests, "1E293B", "FFFFFF", "A5:B6"),
            ("PASSED TESTS", passed_tests, "059669", "FFFFFF", "C5:D6"),
            ("FAILED TESTS", failed_tests, "DC2626" if failed_tests > 0 else "475569", "FFFFFF", "E5:F6"),
            ("PASS RATE", f"{pass_rate:.1f}%", "2563EB", "FFFFFF", "G5:G6")
        ]

        for title, val, bg_color, text_color, cell_range in kpis:
            start_cell = cell_range.split(":")[0]
            ws_summary.merge_cells(cell_range)
            cell = ws_summary[start_cell]
            cell.value = f"{title}\n{val}"
            cell.font = Font(name="Segoe UI", size=13, bold=True, color=text_color)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Executive Summary Section Header
        ws_summary.cell(row=9, column=1, value="Category-Wise Execution Breakdown").font = Font(name="Segoe UI", size=12, bold=True, color="0F172A")

        # Table Header
        cat_headers = ["Category", "Total Tests", "Passed", "Failed", "Pass Rate (%)", "Avg Duration (s)"]
        for col_idx, header in enumerate(cat_headers, start=1):
            c = ws_summary.cell(row=10, column=col_idx, value=header)
            c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Group metrics by category
        categories = {}
        for r in test_results:
            cat = r.get('category', 'General')
            if cat not in categories:
                categories[cat] = {'total': 0, 'pass': 0, 'fail': 0, 'duration': 0.0}
            categories[cat]['total'] += 1
            if r.get('status') == 'PASS':
                categories[cat]['pass'] += 1
            elif r.get('status') == 'FAIL':
                categories[cat]['fail'] += 1
            categories[cat]['duration'] += r.get('duration_sec', 0.0)

        current_row = 11
        for cat, data in categories.items():
            tot = data['total']
            pas = data['pass']
            fal = data['fail']
            crate = (pas / tot * 100) if tot > 0 else 0.0
            avg_dur = (data['duration'] / tot) if tot > 0 else 0.0

            ws_summary.cell(row=current_row, column=1, value=cat).alignment = Alignment(horizontal="left")
            ws_summary.cell(row=current_row, column=2, value=tot).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=current_row, column=3, value=pas).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=current_row, column=4, value=fal).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=current_row, column=5, value=f"{crate:.1f}%").alignment = Alignment(horizontal="center")
            ws_summary.cell(row=current_row, column=6, value=f"{avg_dur:.2f}s").alignment = Alignment(horizontal="center")
            current_row += 1

        # Add Pie Chart for Test Results Breakdown
        try:
            chart = PieChart()
            chart.title = "Test Status Distribution"
            labels = Reference(ws_summary, min_col=1, min_row=11, max_row=current_row - 1)
            data = Reference(ws_summary, min_col=3, min_row=10, max_row=current_row - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 14
            chart.height = 7
            ws_summary.add_chart(chart, "A16")
        except Exception as e:
            logger.warning(f"Could not attach pie chart: {e}")

        # ----------------------------------------------------
        # 2. Detailed Test Results Sheet
        # ----------------------------------------------------
        ws_details = self.wb.create_sheet(title="Test Details & Logs")
        ws_details.views.sheetView[0].showGridLines = True

        headers = [
            "Test ID", "Category", "Role View", "Test Title", "Status",
            "Duration (s)", "Error Details", "Screenshot Path"
        ]

        # Style Header Row
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_details.row_dimensions[1].height = 24

        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        for row_idx, r in enumerate(test_results, start=2):
            status = r.get("status", "UNKNOWN")
            
            # Status colors
            if status == "PASS":
                bg_color = "D1FAE5"  # Light green
                font_color = "065F46"
            elif status == "FAIL":
                bg_color = "FEE2E2"  # Light red
                font_color = "991B1B"
            else:
                bg_color = "FEF3C7"  # Light yellow
                font_color = "92400E"

            ws_details.cell(row=row_idx, column=1, value=r.get("test_id", f"TC-{row_idx-1:03d}")).alignment = Alignment(horizontal="center")
            ws_details.cell(row=row_idx, column=2, value=r.get("category", "General")).alignment = Alignment(horizontal="left")
            ws_details.cell(row=row_idx, column=3, value=r.get("role", "Default")).alignment = Alignment(horizontal="center")
            ws_details.cell(row=row_idx, column=4, value=r.get("title", "")).alignment = Alignment(horizontal="left")

            status_cell = ws_details.cell(row=row_idx, column=5, value=status)
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color=font_color)
            status_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws_details.cell(row=row_idx, column=6, value=f"{r.get('duration_sec', 0.0):.2f}").alignment = Alignment(horizontal="right")
            ws_details.cell(row=row_idx, column=7, value=r.get("error_msg", "-")).alignment = Alignment(horizontal="left")
            ws_details.cell(row=row_idx, column=8, value=r.get("screenshot", "-")).alignment = Alignment(horizontal="left")

            ws_details.row_dimensions[row_idx].height = 20

            # Apply borders
            for c in range(1, len(headers) + 1):
                ws_details.cell(row=row_idx, column=c).border = thin_border

        # Auto-adjust column widths
        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        lines = val_str.split('\n')
                        max_len = max(max_len, max(len(l) for l in lines))
                    else:
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save Workbook
        self.wb.save(filepath)
        logger.info(f"Excel Analysis Report successfully generated at: {filepath}")
        return filepath
